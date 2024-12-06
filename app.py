from flask import Flask, render_template, request, redirect, url_for, flash,send_file
import openpyxl

# Path to your Excel file
excel_load = "static/Doc_supply.xlsx"
excel_modify = None  

app = Flask(__name__)
app.secret_key = 'your_secret_key'

# Mock user credentials for login
VALID_EMAIL = "exemple@gmail.com"
VALID_PASSWORD = "motdepasse1234"

# Temporary storage for form data between steps
gross_requirements = []

def update_excel(quantites_gross, quantites_available):
    global excel_modify
    if not excel_modify:
        raise RuntimeError("Excel modification path not set. Please provide a valid path.")
    
    try:
        wb = openpyxl.load_workbook(excel_modify)
        sheet = wb.active

        # Update Gross Requirements
        ligne_gross = 7
        colonne_depart = 11

        for i, qte in enumerate(quantites_gross):
            colonne = colonne_depart + i
            sheet.cell(row=ligne_gross, column=colonne, value=qte)

        # Update Available (specific rows)
        lignes_available = [9, 16, 23, 30, 37, 44, 51]
        for i, ligne in enumerate(lignes_available):
            if i < len(quantites_available):
                sheet.cell(row=ligne, column=10, value=quantites_available[i])

        wb.save(excel_modify)
        wb.close()
    except Exception as e:
        raise RuntimeError(f"Excel update failed: {e}")

@app.route('/', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        email = request.form.get("email")
        password = request.form.get("password")

        if email == VALID_EMAIL and password == VALID_PASSWORD:
            return redirect(url_for("dashboard"))
        else:
            flash("Invalid email or password. Please try again.", "danger")

    return render_template('login.html')

@app.route('/dashboard')
def dashboard():
    return render_template('dashboard.html')

@app.route('/download-excel')
def download_excel():
    # Serve the Excel file for download
    return send_file(excel_load, as_attachment=True)

@app.route('/update-excel-path', methods=['POST'])
def update_excel_path():
    global excel_modify
    # Get the new path from the form
    new_path = request.form.get("excel_path")

    try:
        # Validate the new path by trying to open the file
        with open(new_path, 'r') as file:
            pass
        excel_modify = new_path
        flash("Excel path for modifications updated successfully!", "success")
    except FileNotFoundError:
        flash("Invalid file path. Please make sure the file exists.", "danger")

    return redirect(url_for('dashboard'))

@app.route('/gross-requirements', methods=['GET', 'POST'])
def gross_requirements():
    global gross_requirements

    if request.method == 'POST':
        try:
            # Store Gross Requirements data and move to next step
            gross_requirements = [int(x) if x else 0 for x in request.form.getlist('gross')]
            return redirect(url_for('available'))
        except Exception as e:
            flash(f"Error while processing Gross Requirements: {e}", "danger")

    return render_template('gross_requirements.html')

@app.route('/available', methods=['GET', 'POST'])
def available():
    global gross_requirements
    categories = ["Skateboard", "Board", "Trucks", "Wheels", "Screws", "Tire", "Rim"]
    if request.method == 'POST':
        try:
            # Process Available data and finalize
            available_quantities = [int(x) if x else 0 for x in request.form.getlist('available')]

            # Update Excel
            update_excel(gross_requirements, available_quantities)
            flash('Excel updated successfully!', 'success')
            return redirect(url_for('confirmation', success=True))
        except Exception as e:
            flash(f"Error while updating Excel: {e}", "danger")
            return redirect(url_for('confirmation', success=False))

    return render_template('available.html',categories=categories)

@app.route('/confirmation')
def confirmation():
    # Check for success or failure based on the query parameter
    success = request.args.get('success', 'false').lower() == 'true'
    message = "Excel updated successfully!" if success else "An error occurred while updating Excel."
    category = "success" if success else "danger"

    return render_template('confirmation.html', message=message, category=category)

if __name__ == '__main__':
    app.run(debug=True)


